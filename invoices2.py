"""
Invoice PDF Data Extractor 
Extrai dados de faturas PDF (número, data, valor) e salva em Excel.
"""

import os
import re
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any
from dataclasses import dataclass, field
from openpyxl import Workbook
import pdfplumber

@dataclass
class Config:
    """Configurações centralizadas"""
    PDF_DIR: Path = Path(r'C:\Users\Felipe\Documents\Projects\python_auto\invoices_pdf')
    EXCEL_DIR: Path = Path(r'C:\Users\Felipe\Documents\Projects\python_auto\excel_sheets')
    HEADERS: Dict[str, str] = field(default_factory=lambda: {
        'A1': 'Número da fatura', 'B1': 'Data de emissão', 'C1': 'Valor da fatura', 
        'D1': 'Nome do arquivo', 'E1': 'Status'
    })


class Patterns:
    """Padrões regex organizados"""
    NUMBER = [r'#\s*(\d+)', r'Invoice\s*#?\s*:?\s*(\d+)', r'Fatura\s*#?\s*:?\s*(\d+)', r'N[úu]mero\s*:?\s*(\d+)']
    DATE = [r'Date\s*:?\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{4})', r'Data\s*:?\s*(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{4})', r'(\d{1,2}[\/\-]\d{1,2}[\/\-]\d{4})']
    VALUE = [r'Total\s*:?\s*R?\$?\s*([\d.,]+)', r'Valor\s*:?\s*R?\$?\s*([\d.,]+)', r'R\$\s*([\d.,]+)', r'(\d{1,3}(?:\.\d{3})*,\d{2})']


@dataclass
class InvoiceData:
    """Dados extraídos de uma fatura"""
    filename: str
    number: Optional[str] = None
    date: Optional[str] = None
    value: Optional[float] = None
    status: str = "Processado"
    
    def to_row(self) -> List[Any]:
        """Converte para linha do Excel"""
        return [
            self.number or "Não encontrado",
            self.date or "Não encontrado", 
            self.value or "Não encontrado",
            self.filename,
            self.status
        ]


class PDFProcessor:
    """Processa PDFs e extrai dados"""
    
    def __init__(self, config: Config = Config()):
        self.config = config
    
    def extract_text(self, pdf_path: Path) -> str:
        """Extrai texto completo do PDF"""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                return "\n".join(page.extract_text() or "" for page in pdf.pages)
        except Exception as e:
            print(f'Erro ao ler {pdf_path.name}: {e}')
            return ""
    
    def find_pattern(self, text: str, patterns: List[str]) -> Optional[str]:
        """Busca primeiro padrão que corresponde"""
        return next((match.group(1) for pattern in patterns 
                    if (match := re.search(pattern, text, re.IGNORECASE))), None)
    
    def clean_value(self, value_str: str) -> Optional[float]:
        """Converte valor monetário brasileiro para float"""
        if not value_str:
            return None
        
        cleaned = re.sub(r'[^\d.,]', '', value_str)
        
        # Formato brasileiro: 1.234,56 -> 1234.56
        if ',' in cleaned and '.' in cleaned:
            cleaned = cleaned.replace('.', '').replace(',', '.')
        elif ',' in cleaned:
            cleaned = cleaned.replace(',', '.')
        
        try:
            return float(cleaned)
        except ValueError:
            return None
    
    def process_pdf(self, pdf_path: Path) -> InvoiceData:
        """Processa um PDF e extrai dados da fatura"""
        invoice = InvoiceData(pdf_path.name)
        
        try:
            text = self.extract_text(pdf_path)
            if not text:
                invoice.status = "Erro: PDF vazio"
                return invoice
            
            invoice.number = self.find_pattern(text, Patterns.NUMBER)
            invoice.date = self.find_pattern(text, Patterns.DATE)
            raw_value = self.find_pattern(text, Patterns.VALUE)
            invoice.value = self.clean_value(raw_value)
            
            # Debug log
            print(f'{pdf_path.name}: Número={invoice.number}, Data={invoice.date}, Valor={invoice.value}')
            
        except Exception as e:
            invoice.status = f"Erro: {e}"
        
        return invoice
    
    def create_excel(self, invoices: List[InvoiceData]) -> Workbook:
        """Cria planilha Excel com os dados"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoices"
            
            # Headers
            for cell, header in self.config.HEADERS.items():
                ws[cell] = header
            
            # Dados
            for i, invoice in enumerate(invoices, start=2):
                for j, value in enumerate(invoice.to_row(), start=1):
                    ws.cell(row=i, column=j, value=value)
        except Exception as e:
            print(f'Erro ao criar Excel: {e}')
            return wb
    
    def run(self) -> bool:
        try: 
            """Executa processamento completo"""
            # Validações
            if not self.config.PDF_DIR.exists():
                print(f'Diretório PDF não encontrado: {self.config.PDF_DIR}')
                return False
            
            self.config.EXCEL_DIR.mkdir(parents=True, exist_ok=True)
            
            # Buscar PDFs
            pdf_files = list(self.config.PDF_DIR.glob('*.pdf'))
            if not pdf_files:
                print('Nenhum PDF encontrado')
                return False
            
            print(f'Processando {len(pdf_files)} PDFs...')
            
            # Processar PDFs
            invoices = [self.process_pdf(pdf) for pdf in pdf_files]
            
            # Gerar Excel
            wb = self.create_excel(invoices)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_path = self.config.EXCEL_DIR / f'invoices_{timestamp}.xlsx'
        except Exception as e:
            print(f'Erro ao processar PDFs: {e}')
            return False
        try:
            wb.save(excel_path)
            print(f'Excel salvo: {excel_path}')
            return True
        except Exception as e:
            print(f'Erro ao salvar: {e}')
            # Fallback
            fallback = Path(f'invoices_{timestamp}.xlsx')
            wb.save(fallback)
            print(f'Salvo no diretório atual: {fallback}')
            return True


def main():
    """Função principal"""
    print("=== Extrator de Faturas PDF ===")
    processor = PDFProcessor()
    success = processor.run()
    print("Concluído!" if success else "Falhou!")


if __name__ == "__main__":
    main()